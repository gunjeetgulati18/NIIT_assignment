###########################################
#Author : Gunjeet Gulati
#Default ASCII encoding
######################################

import requests
import os
import csv
from openpyxl import Workbook
import sys
from time import strptime
from openpyxl import Workbook
from openpyxl import load_workbook
import config
import json
from collections import OrderedDict


class Main(object): 
    '''
        Main class to Handle all the processing for input1 and input2
    '''

    def __init__(self):           
        '''
            initializing dict object
        '''
        self.json_out = OrderedDict()
        
    
    def request_url(self,url):  
        '''
            Made a GET Request on provided URL
        '''
        try:   
            self.requested_url = requests.get(url)
        except Exception as E:
            self.requested_url = 100
        return self.requested_url

    def check_file(self,filename):   
        '''
            Check whether file exist or not 
        '''
        try:       
            file = open(filename, 'r')
        except IOError:
            return False
        return file

    def read_data_csv(self,filename):   
        '''
            Reading data from CSV File
        '''
        try:
            if self.check_file(filename):
                with open(filename,'r') as myFile:
                    last_record = [row for row in csv.reader(myFile)]
                return last_record
        except Exception as E:
            print "line No: ",sys.exc_traceback.tb_lineno
            print E

    def download_file(self,url,filename):           
        '''
            Downloading file using provided url and saving it in provided filename
            ........
            using xlsx as file-type
        '''
        try:

            file_out = self.request_url(url)
            if file_out.status_code == 200:
                if "spreadsheetml" in file_out.headers['Content-Type']:
                    filename = filename+'.xlsx'
                    with open(filename, "wb") as f:
                        f.write(file_out.content)
                    return filename
                else:
                    filename = filename+'.xlsx'
                    with open(filename, "wb") as f:
                        f.write(file_out.content)
                    return filename
        except Exception as E:
            print "line No: ",sys.exc_traceback.tb_lineno
            print E         

    def find_last_record(self,data):      
        '''
            getting last value from column date
        '''   
        try:
            get_last_date = None
            for row,record in enumerate(data[1:]):
                if record[0] !='':
                    get_last_date = record[0]
            return get_last_date
        except:
            return None,None,None,None

    def get_last_record(self,filename):                
        '''
            validating whether file exist and then getting last record value of column date from csv file
        '''
        if self.read_data_csv(filename):
            last_record_for_input1 = self.find_last_record(self.read_data_csv(filename))
        else:
            last_record_for_input1 = None
        return last_record_for_input1


    def get_cell(self,column,row):                  
        '''
            creating cell value on column and row  
        '''
        try:
            get_cell = str(column)+str(row)
            return get_cell
        except Exception as E:
            print E
            print "line No: ",sys.exc_traceback.tb_lineno


    def convert_month_to_value(self,month):         
        '''
            converting month name into month digit
        '''
        try:
            value = strptime(str(month),'%b').tm_mon
        except:
            value = 0
        return value

    def split_date(self,date):           
        '''
            
            spliting  date value into day,month,year using seperator '/'
        '''
        try:
            last_record_mm = date.split('/')[0]
            last_record_dd = date.split('/')[1]
            last_record_yyyy = date.split('/')[2]
        except:
            last_record_mm = 0
            last_record_dd = 0
            last_record_yyyy = 1990
        return last_record_mm,last_record_dd,last_record_yyyy

    def processed_string(self,string):                  
        '''
            remove extra characters from string
        '''
        try:
            clean_string = ''.join(i if i.isalpha() or i =='_' else '' for i in string)
            return clean_string
        except:
            return string

    def get_new_record_value(self,ws,column,row):          
        '''
            get excel cell value on column and row using 'ws' as workbook object
        '''
        return ws[self.get_cell(column,row)].value

    def write_records_in_file(self,output_file,header=None,filetype=None):     
        '''
            writing data in file..Can be csv or json seperated by  'filetype'
            >> 'output_file' as filename 
            >>'header' as data to be generated on file 
            
        '''
        try:
            if filetype == 'JSON':
                myFile = open(output_file, 'w')  
            else:   
                myFile = open(output_file, 'wb')  
            with myFile:
                if filetype =="JSON":
                    myFile.write(json.dumps(header, sort_keys=True, indent=4 ))
                else:
                    writer = csv.writer(myFile)   
                    writer.writerows(header)
                
        except Exception as E:
            print E
            print "line No: ",sys.exc_traceback.tb_lineno

    def append_row_csv(self,output_file,data=None):         
        '''
            append  data in csv file
            >> 'output_file' as filename 
        '''
        try:
            myFile = open(output_file, 'ab')  
            with myFile:
                writer = csv.writer(myFile)   
                writer.writerows(data)
        except Exception as E:
            print E
            print "line No: ",sys.exc_traceback.tb_lineno


    def create_date(self,day,month,year):               
        '''
            creating date from day,month,year
        '''
        return str(month)+'/'+str(day)+'/'+str(year)


    def json_o(self,data,header):
        '''
            converting provided data into json type 
            >>'header' as  key 
            >>'data' as its value
        '''
        try:
            for row,header_name in enumerate(header[0]):
                if header_name in self.json_out:
                    self.json_out[header_name].append(data[row])
                else:
                    self.json_out[header_name] = list([data[row]])
            return self.json_out
        except Exception as E:
            print E
            print "line No: ",sys.exc_traceback.tb_lineno

    
    def output_1(self,ws,file_name,last_record,json_filename=None):        
        '''
            processing records for input1 using workbook object as 'ws' by filtering 
            records through previous processed date saved in reference file as 'last_record'
            >> 'json_filename' provided for json output

        '''
        try:
            self.json_out = OrderedDict()
            row = 12
            last_record_mm,last_record_dd,last_record_yyyy = self.split_date(last_record)
            
            new_year = self.get_new_record_value(ws,'A',row)
            new_month_day = self.get_new_record_value(ws,'B',row)
            
            self.write_records_in_file(file_name,config.input1_header)

            while str(new_year).strip() != 'Memo:':
                if new_year is not None and new_year != "":
                    if int(new_year) >= int(last_record_yyyy):
                        if new_month_day !="" and new_month_day is not None :             
                            if not str(new_month_day).isdigit() :
                                month_value = self.convert_month_to_value(new_month_day)
                                if  month_value > last_record_mm and int(new_year) == int(last_record_yyyy):
                                    new_month = self.convert_month_to_value(new_month_day)
                                elif int(new_year) > int(last_record_yyyy):
                                    new_month = self.convert_month_to_value(new_month_day)
                                else:
                                    new_month = 0
                            else:
                                if new_month_day > last_record_dd and int(new_year) == int(last_record_yyyy) and new_month == last_record_mm :
                                    self.append_row_csv(file_name,[[self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)]])
                                    self.json_o([self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)],config.input1_header)
                                elif  int(new_year) == int(last_record_yyyy) and new_month > last_record_mm :
                                    self.append_row_csv(file_name,[[self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)]])
                                    self.json_o([self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)],config.input1_header)
                                elif  int(new_year) > int(last_record_yyyy) :
                                    self.append_row_csv(file_name,[[self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)]])
                                    self.json_o([self.create_date(new_month_day,new_month,new_year),self.get_new_record_value(ws,'C',row),self.get_new_record_value(ws,'D',row),self.get_new_record_value(ws,'E',row),self.get_new_record_value(ws,'F',row),self.get_new_record_value(ws,'G',row),self.get_new_record_value(ws,'H',row),self.get_new_record_value(ws,'I',row),self.get_new_record_value(ws,'J',row),self.get_new_record_value(ws,'K',row),self.get_new_record_value(ws,'L',row)],config.input1_header)
                            
                row+=1
                new_month_day = ws[self.get_cell('B',row)].value
                
                new_year_ = ws[self.get_cell('A',row)].value 
                if new_year_ is not None and new_year != "" :
                    new_year = ws[self.get_cell('A',row)].value

            self.write_records_in_file(json_filename,self.json_out,'JSON')               
        except Exception as E :
            print E
            print "line No: ",sys.exc_traceback.tb_lineno

    def output_2(self,ws,file_name,last_record,json_filename=None):        
        '''
            processing records for input2 using workbook object as 'ws' by filtering 
            records through previous processed date saved in reference file as 'last_record'
            >> 'json_filename' provided for json output

        '''
        try:
            self.json_out = OrderedDict()
            last_record_mm,last_record_dd,last_record_yyyy = self.split_date(last_record)
            row = 7
            
            new_year = self.get_new_record_value(ws,'A',row)
            new_month_day = self.get_new_record_value(ws,'B',row)
            
            self.write_records_in_file(file_name,config.input2_header)
            
            while not isinstance(new_year,unicode) :
                if new_year is not None and new_year != "":
                    if int(new_year) >= int(last_record_yyyy):
                        if new_month_day !='' and new_month_day is not None :
                            if not str(new_month_day).isdigit() :
                                month_value = self.convert_month_to_value(new_month_day)                                
                                if month_value is not None:
                                    if  month_value > last_record_mm and int(new_year) == int(last_record_yyyy):
                                        self.append_row_csv(file_name,[[self.create_date('1',month_value,new_year),self.get_new_record_value(ws,'C',int(row))]])
                                        self.json_o([self.create_date('1',month_value,new_year),self.get_new_record_value(ws,'C',int(row))],config.input2_header)
                                    elif int(new_year) > int(last_record_yyyy):
                                        self.append_row_csv(file_name,[[self.create_date('1',month_value,new_year),self.get_new_record_value(ws,'C',int(row))]])
                                        self.json_o([self.create_date('1',month_value,new_year),self.get_new_record_value(ws,'C',int(row))],config.input2_header)
                
                row+=1                
                new_month_day = ws[self.get_cell('B',row)].value

                new_year_ = ws[self.get_cell('A',row)].value
                if new_year_ is not None and new_year != "" :
                    new_year = ws[self.get_cell('A',row)].value

            self.write_records_in_file(json_filename,self.json_out,'JSON')       
        except Exception as E :
            print E
            print "line No: ",sys.exc_traceback.tb_lineno


    def main_process_generic(self,input_filename,input_url,output_ref,output_file,option,json_filename=None): 
        '''
            main method to handle both the inputs
            >> 'output_ref' provide reference filename for output ... generated by code itself for finding new update everytime 
            >> 'output_file' provide output filename
            >> 'option' given to select output generation for input1 and  input2
            >> 'json_filename' provided for json output
        '''
        try:
            input_file = self.download_file(input_url,input_filename)                                                    # download input file            
            last_processed_date_for_input = self.get_last_record(output_ref)                                             # get previous record date from reference file
                                                         
            if not last_processed_date_for_input :                                     
                last_processed_date_for_input = self.get_last_record(output_file)                                         # in case of no reference try to get previous record date from last output file
            
            wb = load_workbook(filename = input_file)                                                                     # load workbook for input1
            ws = wb[wb.get_sheet_names()[0]]    
            
            if option == 'INPUT1':                                              
                self.output_1(ws,output_file,last_processed_date_for_input,json_filename)                                   # creating output file for input1
            elif option == 'INPUT2':
                self.output_2(ws,output_file,last_processed_date_for_input,json_filename)                                    # creating output  file for input2
            else:
                return None
            
            last_processed_date_for_input = self.get_last_record(output_file)                                                # get last record date from output1            
            if last_processed_date_for_input:
                self.write_records_in_file(output_ref,[['Last_Output'],[last_processed_date_for_input]])                     # creating reference file for output1 
        
        except Exception as E :
            print E
            print "line No: ",sys.exc_traceback.tb_lineno


class Run_it(Main):   
    '''
        manage execution
    '''
    def __init__(self):     
        '''
            initializing main class object
        '''
        self.setup_object = Main()      

    def start(self):
        '''
            calling method for input1 and input2
        '''
        self.setup_object.main_process_generic(config.input1_filename,config.input1_url,config.output1_ref,config.output1_filename,'INPUT1',config.json_output1_filename)  # creating output1
        self.setup_object.main_process_generic(config.input2_filename,config.input2_url,config.output2_ref,config.output2_filename,'INPUT2',config.json_output2_filename) # creating output2
        


if __name__ == '__main__':
    Run_it().start()